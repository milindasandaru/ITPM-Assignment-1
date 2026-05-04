from playwright.sync_api import sync_playwright
import time
import os
import argparse
import re
from datetime import datetime
from pathlib import Path
import sys
import openpyxl
from openpyxl.cell.cell import MergedCell

# Configuration
ROOT_DIR = Path(__file__).resolve().parent.parent
TESTS_DIR = ROOT_DIR / "test_automation"

DEFAULT_EXCEL_CANDIDATES = [
    str(TESTS_DIR / "Assignment 1 - Test cases.xlsx"),
]

DEFAULT_SHEET_NAME = " Test cases"
DEFAULT_FRONTEND_URL = os.getenv("FRONTEND_URL", "https://www.pixelssuite.com/chat-translator")

DEFAULT_INPUT_COLUMN_CANDIDATES = [
    "Singlish",
    "Input",
    "Singlish Input",
    "Test Input",
    "Source",
    "Sentence",
    "Text",
]

DEFAULT_EXPECTED_COLUMN_CANDIDATES = [
    "Sinhala",
    "Expected_Output",
    "Expected Output",
    "Expected output",
    "Expected",
    "Expected Sinhala",
]

DEFAULT_ACTUAL_COLUMN_CANDIDATES = [
    "Actual_Output",
    "Actual Output",
    "Actual output",
    "Actual",
]

DEFAULT_STATUS_COLUMN_CANDIDATES = [
    "Status",
    "Result",
    "Pass/Fail",
    "Pass Fail",
]

DEFAULT_WAIT_MS = 5000
DEFAULT_RETRIES = 8
DEFAULT_RETRY_WAIT_MS = 1000
DEFAULT_TYPE_DELAY_MS = 30
DEFAULT_TIMEOUT_MS = 60000
DEFAULT_SLOW_MO_MS = 0
DEFAULT_LOG_PREVIEW_CHARS = 70

def _configure_stdout():
    try:
        sys.stdout.reconfigure(encoding="utf-8", errors="backslashreplace")
    except Exception:
        pass
    try:
        sys.stderr.reconfigure(encoding="utf-8", errors="backslashreplace")
    except Exception:
        pass

def _log(level: str, message: str):
    ts = datetime.now().strftime("%H:%M:%S")
    print(f"[{ts}] [{level}] {message}")

def _banner(title: str):
    line = "=" * 72
    print(line)
    print(title)
    print(line)

def _preview_text(value: str, max_chars: int = DEFAULT_LOG_PREVIEW_CHARS) -> str:
    text = "" if value is None else str(value).replace("\n", " ").strip()
    if len(text) <= max_chars:
        return text
    return f"{text[:max_chars]}..."

def _pick_existing_path(candidates):
    for p in candidates:
        if p and os.path.exists(p):
            return p
    return candidates[0] if candidates else None

def _resolve_path(p: str | None) -> str | None:
    if not p:
        return None
    path = Path(p)
    if path.is_absolute():
        return str(path)
    root_candidate = (ROOT_DIR / path).resolve()
    if root_candidate.exists():
        return str(root_candidate)
    tests_candidate = (TESTS_DIR / path).resolve()
    if tests_candidate.exists():
        return str(tests_candidate)
    return str(root_candidate)

def _normalize_header(value) -> str:
    if value is None:
        return ""
    return re.sub(r"[^a-z0-9]+", "", str(value).strip().lower())

def _header_values(ws, row_index: int) -> list:
    max_col = max(1, int(ws.max_column or 1))
    return [ws.cell(row=row_index, column=c).value for c in range(1, max_col + 1)]

def _find_header_row(ws, max_scan_rows: int) -> int:
    input_tokens = {_normalize_header(v) for v in DEFAULT_INPUT_COLUMN_CANDIDATES}
    expected_tokens = {_normalize_header(v) for v in DEFAULT_EXPECTED_COLUMN_CANDIDATES}
    actual_tokens = {_normalize_header(v) for v in DEFAULT_ACTUAL_COLUMN_CANDIDATES}
    status_tokens = {_normalize_header(v) for v in DEFAULT_STATUS_COLUMN_CANDIDATES}

    best_score = -1
    best_row = 1
    scan_limit = max(1, min(int(max_scan_rows), int(ws.max_row or 1)))
    for r in range(1, scan_limit + 1):
        values = _header_values(ws, r)
        texts = [v for v in values if isinstance(v, str) and v.strip() and len(v.strip()) <= 40]
        if len(texts) < 2:
            continue

        norms = {_normalize_header(v) for v in texts}
        if "tcid" in norms and "input" in norms and "expectedoutput" in norms:
            return r

        if "input" not in norms:
            continue
        if not (norms & expected_tokens):
            continue

        score = 0
        for v in texts:
            n = _normalize_header(v)
            if n in input_tokens:
                score += 3
            if n in expected_tokens:
                score += 2
            if n in actual_tokens:
                score += 1
            if n in status_tokens:
                score += 1
        if score > best_score:
            best_score = score
            best_row = r
    return best_row

def _merged_top_left_cell(ws, row: int, col: int):
    cell = ws.cell(row=row, column=col)
    if not isinstance(cell, MergedCell):
        return cell
    for rng in ws.merged_cells.ranges:
        if rng.min_row <= row <= rng.max_row and rng.min_col <= col <= rng.max_col:
            return ws.cell(row=rng.min_row, column=rng.min_col)
    return ws.cell(row=row, column=col)

def _is_top_left_of_merged_cell(ws, row: int, col: int) -> bool:
    cell = ws.cell(row=row, column=col)
    if not isinstance(cell, MergedCell):
        return True
    for rng in ws.merged_cells.ranges:
        if rng.min_row <= row <= rng.max_row and rng.min_col <= col <= rng.max_col:
            return rng.min_row == row and rng.min_col == col
    return True

def _set_cell_value(ws, row: int, col: int, value):
    cell = _merged_top_left_cell(ws, row, col)
    cell.value = value

def _find_column_index(header_values: list, requested_name: str | None, candidates: list[str]) -> int | None:
    indexed = []
    for i, v in enumerate(header_values, start=1):
        if v is None:
            continue
        indexed.append((i, str(v)))

    norm_to_index: dict[str, int] = {}
    for i, v in indexed:
        n = _normalize_header(v)
        if n and n not in norm_to_index:
            norm_to_index[n] = i

    def match(name: str) -> int | None:
        n = _normalize_header(name)
        if not n:
            return None
        if n in norm_to_index:
            return norm_to_index[n]
        for i, v in indexed:
            if n in _normalize_header(v) or _normalize_header(v) in n:
                return i
        return None

    if requested_name:
        found = match(requested_name)
        if found:
            return found

    for c in candidates:
        found = match(c)
        if found:
            return found

    return None

def _last_header_col(header_values: list) -> int:
    last = 0
    for i, v in enumerate(header_values, start=1):
        if v is None:
            continue
        if isinstance(v, str) and not v.strip():
            continue
        last = i
    return last

def _ensure_column(ws, header_row: int, header_values: list, desired_name: str) -> int:
    found = _find_column_index(header_values, desired_name, [])
    if found:
        return found
    col = _last_header_col(header_values) + 1
    ws.cell(row=header_row, column=col).value = desired_name
    if col <= len(header_values):
        header_values[col - 1] = desired_name
    else:
        while len(header_values) < col - 1:
            header_values.append(None)
        header_values.append(desired_name)
    return col

def _dismiss_overlays(page):
    candidates = [
        ("button", re.compile(r"^(Accept|I Agree|Agree|OK|Got it)$", re.IGNORECASE)),
        ("button", re.compile(r"^(Accept all|Accept All)$", re.IGNORECASE)),
    ]
    for role, name in candidates:
        try:
            btn = page.get_by_role(role, name=name).first
            if btn.is_visible():
                btn.click(timeout=2000)
                page.wait_for_timeout(500)
        except Exception:
            pass

def _clear_textarea(page, locator, attempts: int = 3):
    for _ in range(max(1, int(attempts))):
        try:
            locator.click(timeout=2000)
        except Exception:
            pass
        try:
            page.keyboard.press("Control+A")
            page.keyboard.press("Backspace")
        except Exception:
            pass
        try:
            locator.fill("")
        except Exception:
            pass
        try:
            if locator.input_value() == "":
                return
        except Exception:
            pass
        try:
            locator.evaluate(
                """(el) => { el.value = ''; el.dispatchEvent(new Event('input', { bubbles: true })); }"""
            )
            if locator.input_value() == "":
                return
        except Exception:
            pass
        page.wait_for_timeout(200)

def _ensure_input_value(page, input_locator, text: str, type_delay_ms: int):
    _clear_textarea(page, input_locator)
    if type_delay_ms and int(type_delay_ms) > 0:
        input_locator.click(timeout=2000)
        input_locator.type(text, delay=int(type_delay_ms))
    else:
        input_locator.fill(text)
    try:
        current = input_locator.input_value()
        if current is None:
            return
        if str(current).strip() == text.strip():
            return
    except Exception:
        return
    page.wait_for_timeout(150)
    _clear_textarea(page, input_locator)
    input_locator.fill(text)

def _read_output(is_chat: bool, output_locator) -> str:
    if is_chat:
        try:
            v = output_locator.input_value()
            if v is not None:
                v = str(v).strip()
                if v:
                    return v
        except Exception:
            pass
    try:
        v = output_locator.inner_text()
        if v is not None:
            v = str(v).strip()
            if v:
                return v
    except Exception:
        pass
    try:
        v = output_locator.text_content()
        if v is not None:
            v = str(v).strip()
            if v:
                return v
    except Exception:
        pass
    try:
        v = output_locator.evaluate("(el) => el && ('value' in el ? el.value : '')")
        if v is not None:
            v = str(v).strip()
            if v:
                return v
    except Exception:
        pass
    return ""

def _find_chat_locators(page, timeout_ms: int):
    deadline = time.time() + (max(1, timeout_ms) / 1000)
    last_debug = None
    while time.time() < deadline:
        _dismiss_overlays(page)
        try:
            input_by_ph = page.locator('textarea[placeholder*="English"]').first
            output_by_ph = page.locator('textarea[placeholder*="Sinhala"]').first
            if input_by_ph.count() > 0 and output_by_ph.count() > 0 and input_by_ph.is_visible() and output_by_ph.is_visible():
                action = page.get_by_role("button", name=re.compile(r"^Transliterate$", re.IGNORECASE)).first
                return input_by_ph, output_by_ph, action
        except Exception:
            pass

        try:
            count = page.locator("textarea").count()
            visible = []
            for i in range(count):
                loc = page.locator("textarea").nth(i)
                if loc.is_visible():
                    visible.append(loc)
            if len(visible) >= 2:
                action = page.get_by_role("button", name=re.compile(r"^Transliterate$", re.IGNORECASE)).first
                return visible[0], visible[1], action
        except Exception as e:
            last_debug = str(e)

        page.wait_for_timeout(500)

    try:
        meta = page.evaluate(
            """() => Array.from(document.querySelectorAll('textarea')).map(t => ({
              placeholder: t.getAttribute('placeholder') || '',
              disabled: !!t.disabled,
              readOnly: !!t.readOnly,
              visible: !!(t.offsetParent)
            }))"""
        )
        print("Debug: textarea meta:", meta)
    except Exception as e:
        print("Debug: failed to read textarea meta:", e)
    if last_debug:
        print("Debug: last error:", last_debug)
    raise RuntimeError("Could not find Chat UI locators (input/output textareas).")

def _parse_args():
    parser = argparse.ArgumentParser()
    parser.add_argument("--excel", default=_pick_existing_path(DEFAULT_EXCEL_CANDIDATES))
    parser.add_argument("--sheet", default=DEFAULT_SHEET_NAME)
    parser.add_argument("--header-row", type=int, default=0)
    parser.add_argument("--max-header-scan-rows", type=int, default=30)
    parser.add_argument("--input-col", default=None)
    parser.add_argument("--expected-col", default=None)
    parser.add_argument("--actual-col", default=None)
    parser.add_argument("--status-col", default=None)
    parser.add_argument("--url", default=DEFAULT_FRONTEND_URL)
    parser.add_argument("--output", default=None)
    parser.add_argument("--save-every", type=int, default=0)
    parser.add_argument("--headless", action="store_true", default=False)
    parser.add_argument("--wait-ms", type=int, default=DEFAULT_WAIT_MS)
    parser.add_argument("--retries", type=int, default=DEFAULT_RETRIES)
    parser.add_argument("--retry-wait-ms", type=int, default=DEFAULT_RETRY_WAIT_MS)
    parser.add_argument("--type-delay-ms", type=int, default=DEFAULT_TYPE_DELAY_MS)
    parser.add_argument("--timeout-ms", type=int, default=DEFAULT_TIMEOUT_MS)
    parser.add_argument("--slow-mo-ms", type=int, default=DEFAULT_SLOW_MO_MS)
    parser.add_argument("--keep-open", action="store_true", default=False)
    return parser.parse_args()

def run_test():
    _configure_stdout()
    started_at = time.time()
    args = _parse_args()
    args.excel = _resolve_path(args.excel)
    args.output = _resolve_path(args.output) if args.output else args.excel

    _banner("Frontend Transliteration Test Runner")
    _log("INFO", f"Excel file : {args.excel}")
    _log("INFO", f"Sheet      : {args.sheet or '(active sheet)'}")
    _log("INFO", f"URL        : {args.url}")
    _log("INFO", f"Headless   : {args.headless}")
    _log("INFO", f"Output file: {args.output}")

    if not args.excel or not os.path.exists(args.excel):
        _log("ERROR", f"File not found: {args.excel}")
        return

    try:
        wb = openpyxl.load_workbook(args.excel)
    except Exception as e:
        _log("ERROR", f"Error reading Excel file: {e}")
        return

    if args.sheet and args.sheet in wb.sheetnames:
        ws = wb[args.sheet]
    else:
        ws = wb.active

    header_row = int(args.header_row or 0)
    if header_row <= 0:
        header_row = _find_header_row(ws, int(args.max_header_scan_rows))

    header_values = _header_values(ws, header_row)

    input_col_idx = _find_column_index(header_values, args.input_col, DEFAULT_INPUT_COLUMN_CANDIDATES)
    expected_col_idx = _find_column_index(header_values, args.expected_col, DEFAULT_EXPECTED_COLUMN_CANDIDATES)

    if not input_col_idx:
        printable = [str(v) if v is not None else "" for v in header_values]
        _log("ERROR", "Could not resolve input column.")
        _log("INFO", f"Header row: {header_row}")
        _log("INFO", f"Available columns: {printable}")
        return

    actual_col_name = args.actual_col or "Actual output"
    status_col_name = args.status_col or "Status"

    actual_col_idx = _find_column_index(header_values, args.actual_col, DEFAULT_ACTUAL_COLUMN_CANDIDATES)
    status_col_idx = _find_column_index(header_values, args.status_col, DEFAULT_STATUS_COLUMN_CANDIDATES)

    actual_col_idx = actual_col_idx or _ensure_column(ws, header_row, header_values, actual_col_name)
    status_col_idx = status_col_idx or _ensure_column(ws, header_row, header_values, status_col_name)

    rows_total = max(0, int(ws.max_row or 0) - header_row)
    _log("INFO", f"Header row resolved to: {header_row}")
    _log("INFO", f"Input column index      : {input_col_idx}")
    _log("INFO", f"Expected column index   : {expected_col_idx or '(not found)'}")
    _log("INFO", f"Actual column index     : {actual_col_idx}")
    _log("INFO", f"Status column index     : {status_col_idx}")
    _log("INFO", f"Rows to scan            : {rows_total}")

    with sync_playwright() as p:
        # 2. Launch Browser
        if args.headless:
            _log("INFO", "Running in headless mode. Remove --headless to watch typing.")
        browser = p.chromium.launch(headless=args.headless, slow_mo=max(0, int(args.slow_mo_ms)))
        page = browser.new_page()
        page.set_default_timeout(max(1000, int(args.timeout_ms)))

        # 3. Open Frontend
        try:
            page.goto(args.url, wait_until="domcontentloaded")
            try:
                page.wait_for_load_state("networkidle", timeout=max(1000, int(args.timeout_ms)))
            except Exception:
                pass
            page.wait_for_selector("textarea", timeout=max(1000, int(args.timeout_ms)))
            _log("INFO", "Frontend loaded successfully.")
        except Exception as e:
            _log("ERROR", f"Error loading frontend: {e}")
            browser.close()
            return

        is_chat = "chat-translator" in (args.url or "")
        if is_chat:
            try:
                input_locator, output_locator, action_locator = _find_chat_locators(page, int(args.timeout_ms))
            except Exception as e:
                _log("ERROR", f"Error locating chat UI elements: {e}")
                browser.close()
                return
        else:
            input_locator = page.locator("textarea")
            output_locator = page.locator("div.card").filter(has_text=re.compile(r"\\bSinhala\\b")).locator("div.bg-slate-50").first
            action_locator = None

        # 4. Iterate Rows
        processed = 0
        scanned_rows = 0
        skipped_merged_rows = 0
        skipped_empty_rows = 0
        passed_rows = 0
        failed_rows = 0
        collected_rows = 0
        ui_error_rows = 0

        for row_index in range(header_row + 1, int(ws.max_row or 0) + 1):
            scanned_rows += 1
            if not _is_top_left_of_merged_cell(ws, row_index, input_col_idx):
                skipped_merged_rows += 1
                continue

            input_cell = _merged_top_left_cell(ws, row_index, input_col_idx)
            input_value = input_cell.value
            singlish_input = str(input_value).strip() if input_value is not None else ""
            if not singlish_input:
                skipped_empty_rows += 1
                continue

            expected_value = (
                _merged_top_left_cell(ws, row_index, expected_col_idx).value if expected_col_idx else None
            )
            expected_sinhala = str(expected_value).strip() if expected_value is not None else ""

            _log(
                "INFO",
                f"Row {row_index}: input='{_preview_text(singlish_input)}'"
            )

            try:
                _dismiss_overlays(page)
                prev_output = _read_output(is_chat, output_locator)
                _ensure_input_value(page, input_locator, singlish_input, int(args.type_delay_ms))

                if action_locator:
                    action_locator.click()

                page.wait_for_timeout(max(0, int(args.wait_ms)))
                
                # Wait for visible content - retry a few times if empty
                actual_output = ""
                tries = max(1, int(args.retries))
                for i in range(tries):
                    current = _read_output(is_chat, output_locator)
                    if not current:
                        page.wait_for_timeout(max(0, int(args.retry_wait_ms)))
                        continue
                    if prev_output and current == prev_output:
                        page.wait_for_timeout(max(0, int(args.retry_wait_ms)))
                        continue
                    if current:
                        actual_output = current
                        break
                    page.wait_for_timeout(max(0, int(args.retry_wait_ms)))

                if prev_output and actual_output == "" and prev_output != "":
                    raise RuntimeError("Output did not update for this input (still showing previous output).")

                _set_cell_value(ws, row_index, actual_col_idx, actual_output)

                if expected_sinhala:
                    status = "PASS" if actual_output == expected_sinhala else "FAIL"
                else:
                    status = "COLLECTED"
                _set_cell_value(ws, row_index, status_col_idx, status)

                if status == "PASS":
                    passed_rows += 1
                elif status == "FAIL":
                    failed_rows += 1
                else:
                    collected_rows += 1

                _log(
                    "RESULT",
                    f"Row {row_index}: {status} | actual='{_preview_text(actual_output)}'"
                )
                if status == "FAIL":
                    _log(
                        "DETAIL",
                        f"Row {row_index}: expected='{_preview_text(expected_sinhala)}'"
                    )

                processed += 1
                if args.save_every and int(args.save_every) > 0 and processed % int(args.save_every) == 0:
                    wb.save(args.output)
                    _log("INFO", f"Auto-saved workbook at processed rows: {processed}")
                
            except Exception as e:
                ui_error_rows += 1
                _log("ERROR", f"Row {row_index}: UI interaction failed: {e}")
                try:
                    _set_cell_value(ws, row_index, status_col_idx, "UI Error")
                except Exception:
                    pass
                if args.save_every and int(args.save_every) > 0:
                    try:
                        wb.save(args.output)
                        _log("WARN", "Workbook saved after UI error.")
                    except Exception:
                        pass

        if args.keep_open and not args.headless:
            try:
                wb.save(args.output)
            except Exception:
                pass
            _log("INFO", "Keeping browser open. Press CTRL+C to stop.")
            try:
                while True:
                    page.wait_for_timeout(1000)
            except KeyboardInterrupt:
                try:
                    wb.save(args.output)
                except Exception:
                    pass
        browser.close()

    try:
        wb.save(args.output)
    except Exception as e:
        _log("ERROR", f"Error saving output file '{args.output}': {e}")
        return

    duration = time.time() - started_at
    _banner("Execution Summary")
    _log("INFO", f"Rows scanned         : {scanned_rows}")
    _log("INFO", f"Rows processed       : {processed}")
    _log("INFO", f"Rows skipped (merged): {skipped_merged_rows}")
    _log("INFO", f"Rows skipped (empty) : {skipped_empty_rows}")
    _log("INFO", f"PASS                 : {passed_rows}")
    _log("INFO", f"FAIL                 : {failed_rows}")
    _log("INFO", f"COLLECTED            : {collected_rows}")
    _log("INFO", f"UI Error             : {ui_error_rows}")
    _log("INFO", f"Duration             : {duration:.2f}s")
    _log("INFO", f"Results saved to     : {args.output}")

if __name__ == "__main__":
    run_test()
